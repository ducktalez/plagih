import copy
import re
from pathlib import Path
import xmltodict

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.merge import MergeCells


class Exls:
    yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    fill_error = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    fill_ifend2 = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
    fill_ifend3 = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='thick'), right=Side(style='thick'),
                          top=Side(style='thick'), bottom=Side(style='thick'))
    merge_a_to_c = lambda xx: MergeCells(start_row=xx, start_column=1, end_row=xx, end_column=4)


class ExcelRowBase:

    def __init__(self, columns: iter, style: 'dict' = None):
        self.columns = columns
        self.style = style

    # def __repr__(self):
    #     return f'{self.columns} {self.style}'

    def __str__(self):
        prs = self.style or ''
        try:
            vs = list(self.columns.values())
            vs = [x for x in vs if x is not None]
            vs = '\t'.join(vs)
        except TypeError as sfeh:
            vs = self.columns
        return f'{vs} \t\t{prs}'


# class Excel_package:
#
#     def __init__(self, name, description, rows=None):
#         self.rows = [ExcelRowBase({2: name}, style={'TODO'}),
#                      ExcelRowBase({2: description}, style='TODO')]
#
#     def add_teststep(self, stepdict):
#         dc = {1: stepdict['Aktion'],
#               2: stepdict['Variablenname'],
#               3: stepdict['Vorgabe/Erwartungswert']}
#         style = None
#         row = ExcelRowBase(dc, style)
#         self.rows.append(row)
#
#     def add_testblock(self, name):
#         row = (ExcelRow_Teststep({2: name}, style=(ExcelStyle.yellow_fill, ExcelStyle.merge_a_to_c)))
#         self.rows.append(row)


def test_triple(stepdict, style):
    dc = {1: stepdict['Aktion'],
          2: stepdict['Variablenname'],
          3: stepdict['Vorgabe/Erwartungswert']}
    # if style is None:  todo
    style['row_boarder'] = style.get('row_boarder') or Exls.thin_border
    row = ExcelRowBase(dc, style)
    return row


class Excel_Sheet:

    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.rows = []

    def append(self, row: ExcelRowBase):
        self.rows.append(row)

    def extend(self, rows: iter('ExcelRowBase')):
        self.rows.extend(rows)


def excel_row_border(row, style='thin'):
    row[0].border = Border(left=Side(style=style), top=Side(style=style), bottom=Side(style=style))
    row[-1].border = Border(right=Side(style=style), top=Side(style=style), bottom=Side(style=style))
    for cell in row[1:-1]:
        cell.border = Border(top=Side(style=style), bottom=Side(style=style))


def get_VALUE_expr(value):
    ex_type = value['@xsi:type']
    if ex_type == 'valueBaseExpression':
        b1 = value['VALUE']['@xsi:type']
        b2 = value['VALUE']['#text']
        b = f'{b2}'
    elif ex_type == 'expressionValue':
        data = value['DATA']
        b = get_VALUE_expr(data)
    elif ex_type == 'varBaseExpression':
        b = value['NAME']['#text']

    elif ex_type == 'binaryOpBaseExpression' or ex_type == 'compareOpBaseExpression':
        def get_component_string(c):
            if c['@xsi:type'] == 'varBaseExpression':
                return c['NAME']['#text']
            elif c['@xsi:type'] == 'valueBaseExpression':
                return c['VALUE']['#text']

        b1 = value['@xsi:type']
        b2 = value['NAME']['#text']
        b2 = re.sub('BINARY_ADD', '+', b2)
        b2 = re.sub('BINARY_SUB', '-', b2)
        b3 = get_component_string(value['FIRST-COMPONENT'])
        b4 = get_component_string(value['SECOND-COMPONENT'])
        b = f'({b3} {b2} {b4})'

    elif ex_type == 'unaryOpBaseExpression':
        b2 = value['NAME']['#text']
        b2 = re.sub('UNARY_NEGATIVE', '-', b2)
        c = value['COMPONENT']
        c = get_VALUE_expr(c)
        b = f'{b2}{c}'

    elif ex_type == 'listGroupOption' or ex_type == 'listExistOption':
        b = get_EXPRESSION(value)  # todo

    else:
        raise NotImplementedError

    return b


def get_TOLERANCE(expression):
    tolerance = expression.get('TOLERANCE')
    if tolerance is not None:
        a = tolerance['@style']  # percentage
        a = re.sub('percentage', '%', a)
        a = re.sub('absolute-value', '', a)
        b = tolerance['VALUE']['VALUE']['#text']
        s = f' [+-{b}{a}]'
    else:
        s = ''

    return s


def get_EXPRESSION(expectation):
    expression = expectation.get('EXPRESSION')
    if expression is not None:
        xsi_type = expression['@xsi:type']
        if xsi_type == 'builtNumericExpression':
            rel = expression['RELATION']['#text']
            value = expression['VALUE']
            b = get_VALUE_expr(value)
            tolerance = get_TOLERANCE(expression)
            s = f'{rel} {b}{tolerance}'
        elif xsi_type == 'builtStringExpression':
            s = expression['BASE-EXPRESSION']['VALUE']['#text']
            s = f'{s}'
        else:
            raise NotImplementedError

        return s


def get_EXPECTATION(x):
    expectation = x.get('EXPECTATION')
    if expectation is not None:

        a = get_EXPRESSION(expectation)
        if expectation.get('@xsi:type') == 'timelessOption':
            s_time = ''
        elif expectation.get('@xsi:type') == 'finallyTrueOption' or expectation.get(
                '@xsi:type') == 'generallyTrueOption':
            e_time = expectation['TIME']['VALUE']['#text']
            try:
                t_unit = expectation['TIME-UNIT']['#text']  # only occurs in 'generallyTrueOption'?
            except Exception as ex:
                t_unit = 'ms'

            s_time = f'  [t: {e_time}{t_unit}]'
        else:
            raise NotImplementedError

        return f' {a}{s_time}'
    else:
        return f''


def get_SAVETO(x):
    s = x.get('VARIABLE-REFS')
    if s is not None:
        s = s['VARIABLE-NAME']['DVALUE']['#text']
        return f' -> {s}'
    else:
        return ''


def get_METRIC_expr(metric):
    try:
        s = metric['Z-UNIT']['#text']
        s = f' {s}'
    except KeyError as ex:
        s = ''
    s = re.sub('u_none', '', s)
    return s


def sanitize(ss):
    ss = re.sub('DME1_DDE1/', '', ss)
    ss = re.sub('/Control/Value', '', ss)
    ss = re.sub('/Active/Value', '', ss)
    # Now, remove the path start? Rossmann 'BK_AE_MOTBK_2010/ENGDAT_1_MOTBK_2010/ENGDAT_1_MOTBK_2010/DISP_MIL_MOTBK_2010'
    ss = re.sub(r'.*\/', '', ss)
    return ss


IGNORE_USELESS_LEVEL = 1


def teststep_get_excel(ecutest):
    if (ecutest.get('@xsi:type') or '') == 'noEventTestStepMappingContainer':
        return

    elif (ecutest.get('@xsi:type') or '') == 'testCase':
        return

    elif (ecutest.get('#text') or '') == '<padding>':
        return

    elif (ecutest.get('@name') or '') == 'TsBlock':  # @xsi:type is 'utility-2752ad1e-4fef-11dc-81d4-0013728784ee'
        bb = ecutest['ACTION']['MULTILANGDATA']['ELEMENT']['DVALUE']['#text']
        if IGNORE_USELESS_LEVEL >= 1 and ('Zuruecksetzen' in bb):
            return
        elif IGNORE_USELESS_LEVEL >= 2 and ('Precondition' in bb or 'Postcondition' in bb):
            return
        else:
            EXCEL_dict = test_triple({'Aktion': '', 'Variablenname': bb, 'Vorgabe/Erwartungswert': f''},
                                     style={'fill': Exls.yellow_fill, 'row_border': Exls.thin_border,
                                            'is_teststep': True})

    elif (ecutest.get('@xsi:type') or '') == 'tsJob':
        s = ecutest['MAPPING-REF']['#text']
        EXCEL_dict = test_triple({'Aktion': 'tsJob', 'Variablenname': f'{s}', 'Vorgabe/Erwartungswert': f''},
                                 style={'row_border': Exls.thin_border, 'is_teststep': True})

    elif (ecutest.get('@xsi:type') or '') == 'tsPackage':
        try:
            ecupath = ecutest['PACKAGE-REFERENCE']['VALUE']['#text']
        except KeyError:
            ecupath = ecutest['PACKAGE-REFERENCE']['PATH-EXPRESSION']['VALUE']['#text']

        ecupath = re.sub(r'.*\\', '', ecupath)

        if ecutest.get('PARAM-ASSIGNMENTS') is not None:
            params = ecutest['PARAM-ASSIGNMENTS']['ASSIGNMENT']
            b = []
            if isinstance(params, list):
                for p in params:
                    p0 = p['@dkey']
                    p1 = get_VALUE_expr(p['DVALUE'])
                    b.append(f'{p0}={p1}')
            elif isinstance(params, dict):
                p0 = params['@dkey']
                p1 = get_VALUE_expr(params['DVALUE'])
                b.append(f'{p0}={p1}')
            else:
                raise NotImplementedError
            b = ', '.join(b)
            b = f' {b}'
            if IGNORE_USELESS_LEVEL >= 2 and ('HIL Init.pkg' in ecupath or 'TerminateExecution.pkg' in ecupath):
                return
            elif IGNORE_USELESS_LEVEL >= 1 and ('HIL Init.pkg' in ecupath or 'TerminateExecution.pkg' in ecupath):
                b = ''
        else:
            b = ''
        EXCEL_dict = test_triple({'Aktion': 'Aufruf', 'Variablenname': ecupath, 'Vorgabe/Erwartungswert': f'{b}'},
                                 style={'fill': None, 'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@xsi:type') == 'tsEesError':
        a = ecutest['EES-ERROR']['PINS']['ELEMENT']['NAME']['#text']
        b = ecutest['EES-ERROR']['PINS']['ELEMENT']['LOAD']['#text']
        c = ecutest['CONTAINED-MAPPING-ITEMS']['CONTAINED-MAPPING-ITEM']
        # Rossmann: Welche Infos?
        EXCEL_dict = test_triple({'Aktion': 'Error', 'Variablenname': f'{c}', 'Vorgabe/Erwartungswert': f''},
                                 style={'fill': None, 'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@xsi:type') == 'tsRead':
        ecupath = ecutest['MAPPING-REF']['#text']
        if IGNORE_USELESS_LEVEL >= 1 and ('/Active/Value' in ecupath or ecupath[-7:] == '_Switch'):
            return
        ecupath = sanitize(ecupath)
        save_to = get_SAVETO(ecutest)
        expectation = get_EXPECTATION(ecutest)
        # s = f'{ecupath}{expectation}{save_to}'
        if expectation is not None:
            aa = 'prüfen'
        elif save_to is not None:
            aa = 'speichern'
        else:
            raise NotImplementedError('What is done here??')
        EXCEL_dict = test_triple(
            {'Aktion': aa, 'Variablenname': ecupath, 'Vorgabe/Erwartungswert': f'{expectation}{save_to}'},
            style={'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@xsi:type') == 'tsDiagEdiabas':
        # Rossmann: was von Ediabas?
        b = ecutest['MAPPING-REF']['#text']  # 'DME1_DDE1/FS_LESEN'
        b = re.sub('DME1_DDE1/', '', b)
        try:
            ar = ecutest['ARGUMENTS']['ARGUMENT']
            if isinstance(ar, list):
                acc = []
                for x in ar:
                    x0 = x['NAME']['#text']
                    x1 = get_VALUE_expr(x['VALUE'])
                    acc.append(f'{x0}={x1}')
                c = ', '.join(acc)
            else:

                c1 = ecutest['ARGUMENTS']['ARGUMENT']['NAME']['#text']  # 'FEHLER_KLASSE'
                c2 = ecutest['ARGUMENTS']['ARGUMENT']['VALUE']['VALUE']['#text']  # '0x3a000D'
                c = f'{c1}={c2}'
        except TypeError as ex:
            c = ''  # FS_LOESCHEN

        try:
            options = ecutest['EXPECTED-RESULTS']['OPTION']
            acc = []
            for oo in ecutest['EXPECTED-RESULTS']['OPTION']:
                dkey = oo['@dkey']
                dvalue = oo['DVALUE']
                dvalue = get_VALUE_expr(dvalue)
                acc.append(f'{dkey}({dvalue})')
            acc = ', '.join(acc)
            cd = f'args: ({c}) expected-results: ({acc})'
        except Exception as ex:
            cd = f'args: ({c})'

        # BK_AE_MOTBK_2010/ENGDAT_1_MOTBK_2010/ENGDAT_1_MOTBK_2010/DISP_MIL_MOTBK_2010
        EXCEL_dict = test_triple({'Aktion': 'Ediabas', 'Variablenname': b, 'Vorgabe/Erwartungswert': cd},
                                 style={'fill': None, 'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@xsi:type') == 'tsWrite':
        ecupath = ecutest['MAPPING-REF']['#text']
        if IGNORE_USELESS_LEVEL >= 1 and ('/Active/Value' in ecupath or ecupath[-7:] == '_Switch'):
            return
        ecupath = sanitize(ecupath)
        value = ecutest['VALUE']
        b = get_VALUE_expr(value)
        EXCEL_dict = test_triple({'Aktion': 'schreiben', 'Variablenname': ecupath, 'Vorgabe/Erwartungswert': f'{b}'},
                                 style={'fill': None, 'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@xsi:type') == 'tsRestore':
        # IGNORE_USELESS_LEVEL = 2
        return

    elif ecutest.get('@xsi:type') == 'ifThenElseNode':
        return

    elif ecutest.get('@name') == 'TsWait':
        # s = ecutest['TIME']['VALUE']['#text']
        # s = f'--wait-- {s}'
        # IGNORE_USELESS_LEVEL = 3?
        return

    elif ecutest.get('@name') == 'TsComment':
        return

    elif ecutest.get('@name') == 'TsEesErrorSet':
        b = ecutest['NAME']['#text']  # Rossmann
        EXCEL_dict = test_triple({'Aktion': 'TsEesErrorSet', 'Variablenname': f'{b}', 'Vorgabe/Erwartungswert': ''},
                                 style={'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@name') == 'TsIfThenElse':
        # Rossmann: if statements?
        cond = ecutest['CONDITION']
        cond = get_VALUE_expr(cond)
        EXCEL_dict = test_triple({'Aktion': 'IF statement', 'Variablenname': f'{cond}', 'Vorgabe/Erwartungswert': ''},
                                 style={'row_border': Exls.thin_border, 'is_teststep': True, 'fill': Exls.fill_error})

    elif ecutest.get('@name') == 'TsImageUI':
        b = ecutest['DLG-TITLE']
        b = get_VALUE_expr(b)
        EXCEL_dict = test_triple({'Aktion': 'USER_INPUT', 'Variablenname': f'{b}', 'Vorgabe/Erwartungswert': ''},
                                 style={'row_border': Exls.thin_border, 'is_teststep': True, 'fill': Exls.fill_error})

    elif ecutest.get('@name') == 'TsLoop':
        loopcount = ecutest['LOOP-COUNT']
        loopcount = get_VALUE_expr(loopcount)  # Rossmann: loops?
        EXCEL_dict = test_triple({'Aktion': 'TsLoop', 'Variablenname': f'LOOP', 'Vorgabe/Erwartungswert': loopcount},
                                 style={'row_border': Exls.thin_border, 'is_teststep': True, 'fill': Exls.fill_error})

    elif ecutest.get('@name') == 'TsCalculation':
        s = ecutest['FORMULA']  # todo
        s = get_VALUE_expr(s)
        EXCEL_dict = test_triple({'Aktion': 'TsCalculation', 'Variablenname': f'{s}', 'Vorgabe/Erwartungswert': None},
                                 style={'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@name') == 'TsStartTrace':
        s = ecutest['NAME']['#text']
        EXCEL_dict = test_triple({'Aktion': 'Traceanalyse (start)', 'Variablenname': s, 'Vorgabe/Erwartungswert': None},
                                 style={'fill': None, 'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@name') == 'TsStopTrace':
        s = ecutest['NAME']['#text']
        EXCEL_dict = test_triple(
            {'Aktion': 'Traceanalyse (stop)', 'Variablenname': f'{s}', 'Vorgabe/Erwartungswert': None},
            style={'row_border': Exls.thin_border, 'is_teststep': True})

    elif ecutest.get('@xsi:type') == 'list':
        return
    else:
        raise NotImplementedError

    return EXCEL_dict


def ecu_to_excel_recursive(ecu_xml, depth=0):
    if ecu_xml.get('ENABLED') or False:
        # print(f'Disabled!')
        return
    else:
        try:
            ecu_xml = ecu_xml['TESTSTEPS']
        except KeyError:
            pass

        teststep_xml = ecu_xml.get('TESTSTEP') or None

        if teststep_xml is None:  # None, if no teststeps
            excel_teststep = teststep_get_excel(ecu_xml)
            if ecu_xml.get('@name') == 'TsIfThenElse':
                a = ecu_xml['THEN']
                a = ecu_to_excel_recursive(a, depth=depth+1)
                for x in a:
                    x.style['fill'] = Exls.fill_ifend2
                b = ecu_xml['ELSE']
                # Rossmann: ignore else-case
                """
                b = ecu_to_excel_recursive(b)
                """
                try:
                    for x in b:
                        x.style['fill'] = Exls.fill_ifend3
                except Exception as ex:
                    b = ExcelRowBase

                excel_teststep = [excel_teststep] + a + b  # Rossmann how to handle if-statements?
                return excel_teststep
            else:
                return [excel_teststep]

        else:
            if isinstance(teststep_xml, dict):  # dict, if only one teststep
                excel_teststep = teststep_get_excel(teststep_xml)
                return [excel_teststep]

            elif isinstance(teststep_xml, list):  # list, if multiple teststep
                try:
                    excel_list = [teststep_get_excel(ecu_xml)]
                except Exception as todo:
                    excel_list = []
                for tt in teststep_xml:
                    tn = ecu_to_excel_recursive(tt, depth=depth+1)
                    if isinstance(tn, list) and len(tn) > 0:
                        excel_list.extend(tn)
                    elif isinstance(tn, ExcelRowBase):
                        excel_list.append(tn)
                    elif tn is None:
                        pass  # e.g. not 'ENABLED', irrelevant teststep, ...
                    else:
                        raise NotImplementedError
                # excel_list = [x[0] if isinstance(x, list) else x for x in excel_list]
                excel_list = [x for x in excel_list if x is not None]
                return excel_list
            else:
                raise NotImplementedError


def xls_to_excel_writer(rows):
    excel_path = 'example.xlsx'
    excel_sheet = 'Sheet1'
    row_counter = 0

    # DataFrame in eine Excel-Datei speichern
    # Die Excel-Datei mit openpyxl laden
    workbook = Workbook()
    ws = workbook.create_sheet(excel_sheet, 0)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    workbook.active = workbook[excel_sheet]
    # ws['A'].alignment = Alignment(wrap_text=True)

    # # Überprüfen und Zeilen färben, wenn in Reihe 2 der Text "Block" steht
    # for row in ws.iter_rows(min_row=0, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    #     print(f'{row[0].value}: {row[0]}')
    #     if row[0].value is None:  # Block, Index 1 bezieht sich auf die zweite Spalte (B)
    #         for cell in row:
    #             cell.fill = ExcelStyle.yellow_fill
    #         excel_row_border(row)
    for ii, row in enumerate(rows):
        ii = ii + 1  # rows start with 1 >.>

        try:
            vs = list(row.columns.values())
        except Exception as ex:
            continue  # None... idk why
        style = row.style
        ws.append(vs)

        try:
            if style.get('alignment') == True:
                ws.merge_cells(start_row=ii, end_row=ii, start_column=1, end_column=3)
                c = ws.cell(column=1, row=ii)
                c.alignment = Alignment(wrap_text=True)
                rd = ws.row_dimensions[ii]
                rd.height = 15 + vs[0].count('\n') * 15  # {'defaultRowHeight': '15', 'baseColWidth': '10'}

            if style.get('row_border'):
                for col in [1, 2, 3]:
                    c = ws.cell(column=col, row=ii)
                    asd = style.get('row_border')
                    c.border = asd

            if style.get('bold'):
                for col in [1, 2, 3]:
                    c = ws.cell(column=col, row=ii)
                    c.font = Font(bold=True)

            if style.get('fill'):
                fill = style.get('fill')
                for col in [1, 2, 3]:
                    c = ws.cell(column=col, row=ii)
                    c.fill = fill

        except Exception as ex:
            pass

    # Änderungen speichern
    workbook.save(excel_path)

    print(f"Excel-Datei wurde erfolgreich unter {excel_path} gespeichert und formatiert.")


# def xls_to_excel_writer
#
#     with pd.ExcelWriter(excel_path) as writer:
#         for row in rows:
#             row.ecu_to_excel_recursive(writer, engine="xlsxwriter", sheet_name='Sheet1', startrow=startrow, index=False)
#             startrow += (.shape[0] + 3)

PATH_PROJECTS = Path('C:/Users/Simon/Documents/BMW-Motorrad/#Oktober/HIL_AE/Packages/Projects/')
PATH_PACKAGES = Path('C:/Users/Simon/Documents/BMW-Motorrad/#Oktober/HIL_AE/Packages/')


def proj_get_testcase(testcase):
    if testcase['ENABLED']['#text'] == 'True':
        if testcase['@xsi:type'] == 'subproject':
            raise NotImplementedError
        elif testcase['@xsi:type'] == 'packageTest':
            if testcase['ENABLED']['#text'] == 'True':
                p_testcase = PATH_PROJECTS / testcase['PACKAGE-REF']['PACKAGE-PATH']['#text']  # creates actual path
                if not p_testcase.exists():
                    raise FileNotFoundError(f'Testcase not found: {p_testcase}')
                return p_testcase
        else:
            raise NotImplementedError
    else:
        return None


def get_testcase_dict():
    # C:\Users\Simon\Documents\BMW-Motorrad\#Oktober\HIL_AE\Packages\Projects\AlleDTCs.prj
    # AlleDTCs
    sheet_cases_dict = {}
    with Path.open('C:/Users/Simon/Documents/BMW-Motorrad/#Oktober/HIL_AE/Packages/Projects/AlleDTCs.prj', 'r',
                   encoding='utf-8') as file:
        data = file.read()
        proj = xmltodict.parse(data)
    subname = proj['PROJECT']['COMPONENTS']['COMPONENT'][3]['NAME']['#text']

    sheet_cases_dict['OBD_DTCs_BMW'] = []
    for tt in proj['PROJECT']['COMPONENTS']['COMPONENT'][3]['COMPONENTS']['COMPONENT'][0]['COMPONENTS']['COMPONENT']:
        sheet_cases_dict['OBD_DTCs_BMW'].append(proj_get_testcase(tt))
    # todo
    # sheet_cases_dict['OBD_DTCs_Marelli'] = []
    # for tt in proj['PROJECT']['COMPONENTS']['COMPONENT'][3]['COMPONENTS']['COMPONENT'][1]['COMPONENTS']['COMPONENT']:
    #     sheet_cases_dict['OBD_DTCs_Marelli'].append(proj_get_testcase(tt))
    #
    # sheet_cases_dict['Other'] = []
    # for tt in proj['PROJECT']['COMPONENTS']['COMPONENT'][3]['COMPONENTS']['COMPONENT'][2:]:
    #     # name = tt['NAME']['#text']
    #     sheet_cases_dict['Other'].append(proj_get_testcase(tt))

    return sheet_cases_dict


TESTCASE_INIT = 1


def xls_from_package_file(file: Path):
    if file is not None:

        with file.open('r', encoding='utf-8') as f:
            data = f.read()
            ecu_testcase_xml = xmltodict.parse(data)

        try:
            p_desc = ecu_testcase_xml['PACKAGE']['INFORMATION']['DESCRIPTION']['#text']
        except KeyError:
            p_desc = None

        xls_rows = [ExcelRowBase({2: file.name}, style={'row_border': Exls.thin_border}),
                    ExcelRowBase({2: p_desc}, style={'row_border': Exls.thin_border, 'alignment': True}),
                    ExcelRowBase({1: 'Aktion', 2: 'Variablenname', 3: 'Vorgabe/Erwartungswert'},
                                 style={'bold': True, 'row_border': Exls.thick_border})]

        testpkg = ecu_testcase_xml['PACKAGE']
        xls_subrows = ecu_to_excel_recursive(testpkg, depth=0)
        xls_rows.extend(xls_subrows)

        placeholder_todo = ExcelRowBase({0: ''})
        placeholder1_todo = ExcelRowBase({0: '====================================================================='})
        xls_rows.append(placeholder1_todo)  # just clear some lines
        xls_rows.append(placeholder_todo)  # just clear some lines
        return xls_rows
    else:
        placeholder1_todo = ExcelRowBase({0: '====================================================================='})
        xls_rows = [ExcelRowBase({2: 'SKIPPED'}, style='TODO'),
                    placeholder1_todo]
        return xls_rows  # todo


subproj_packlist_dict = get_testcase_dict()
xlssheets_xlsdict = []
for package_p in subproj_packlist_dict['OBD_DTCs_BMW']:
    xls_rows = xls_from_package_file(package_p)

    xlssheets_xlsdict.extend(xls_rows)
xls_to_excel_writer(xlssheets_xlsdict)
